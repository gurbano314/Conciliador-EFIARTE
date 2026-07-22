"""
main_qt.py — Revisor de Informes Semestrales EFIARTES
Aplicación de escritorio PyQt6 para revisión normativa de informes EFIARTES.
"""
import sys
import os
import json
import math
import base64
from typing import Optional

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QHBoxLayout, QVBoxLayout,
    QPushButton, QLabel, QScrollArea, QFileDialog, QSplitter,
    QListWidget, QListWidgetItem, QLineEdit, QTextEdit, QComboBox,
    QCheckBox, QFrame, QProgressBar, QStatusBar, QMessageBox,
    QFormLayout, QSizePolicy, QToolButton, QGroupBox, QSpinBox,
)
from PyQt6.QtCore import (
    Qt, QThread, pyqtSignal, QSize, QTimer, QRectF,
)
from PyQt6.QtGui import (
    QColor, QFont, QIcon, QPainter, QPixmap, QImage, QPen, QBrush,
    QFontDatabase,
)

import engine

# ─────────────────────────────────────────────────────────────
# PALETA DE COLORES (Azul marino SHCP)
# ─────────────────────────────────────────────────────────────
C_BG       = "#0D1117"
C_PANEL    = "#161B22"
C_SIDEBAR  = "#1C2128"
C_BORDER   = "#2D3748"
C_ACCENT   = "#1A6FC4"
C_ACCENT2  = "#2589E0"
C_TEXT     = "#E6EDF3"
C_MUTED    = "#8B949E"
C_OK       = "#3FB950"
C_WARN     = "#D29922"
C_FAIL     = "#F85149"
C_MANUAL   = "#8B949E"
C_REVIEWED = "#1A6FC4"

GLOBAL_STYLE = f"""
QWidget {{
    background-color: {C_BG};
    color: {C_TEXT};
    font-family: 'Segoe UI', Arial, sans-serif;
    font-size: 13px;
}}
QScrollBar:vertical {{
    background: {C_PANEL};
    width: 8px;
    border-radius: 4px;
}}
QScrollBar::handle:vertical {{
    background: {C_BORDER};
    border-radius: 4px;
    min-height: 20px;
}}
QScrollBar::handle:vertical:hover {{
    background: {C_ACCENT};
}}
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
    height: 0px;
}}
QScrollBar:horizontal {{
    background: {C_PANEL};
    height: 8px;
    border-radius: 4px;
}}
QScrollBar::handle:horizontal {{
    background: {C_BORDER};
    border-radius: 4px;
    min-width: 20px;
}}
QListWidget {{
    background: {C_PANEL};
    border: 1px solid {C_BORDER};
    border-radius: 6px;
    outline: none;
    padding: 4px;
}}
QListWidget::item {{
    border-radius: 4px;
    padding: 4px 6px;
    margin: 1px 0;
}}
QListWidget::item:selected {{
    background: {C_ACCENT};
    color: white;
}}
QListWidget::item:hover:!selected {{
    background: {C_SIDEBAR};
}}
QPushButton {{
    background-color: {C_ACCENT};
    color: white;
    border: none;
    border-radius: 6px;
    padding: 7px 14px;
    font-weight: 600;
    font-size: 12px;
}}
QPushButton:hover {{
    background-color: {C_ACCENT2};
}}
QPushButton:disabled {{
    background-color: {C_BORDER};
    color: {C_MUTED};
}}
QPushButton#btn_secondary {{
    background-color: {C_SIDEBAR};
    border: 1px solid {C_BORDER};
    color: {C_TEXT};
}}
QPushButton#btn_secondary:hover {{
    background-color: {C_BORDER};
}}
QPushButton#btn_danger {{
    background-color: #6E1A1A;
    border: 1px solid #C0392B;
    color: {C_TEXT};
}}
QPushButton#btn_danger:hover {{
    background-color: #8B2020;
}}
QPushButton#btn_ok {{
    background-color: #1A4A28;
    border: 1px solid {C_OK};
    color: {C_TEXT};
}}
QPushButton#btn_ok:hover {{
    background-color: #256336;
}}
QLineEdit, QTextEdit, QComboBox, QSpinBox {{
    background: {C_PANEL};
    border: 1px solid {C_BORDER};
    border-radius: 5px;
    padding: 5px 8px;
    color: {C_TEXT};
    selection-background-color: {C_ACCENT};
}}
QLineEdit:focus, QTextEdit:focus {{
    border: 1px solid {C_ACCENT};
}}
QComboBox::drop-down {{
    border: none;
    width: 20px;
}}
QComboBox::down-arrow {{
    image: none;
    border-left: 4px solid transparent;
    border-right: 4px solid transparent;
    border-top: 5px solid {C_MUTED};
    margin-right: 5px;
}}
QComboBox QAbstractItemView {{
    background: {C_PANEL};
    border: 1px solid {C_BORDER};
    selection-background-color: {C_ACCENT};
    outline: none;
}}
QGroupBox {{
    border: 1px solid {C_BORDER};
    border-radius: 6px;
    margin-top: 10px;
    padding-top: 10px;
    font-weight: 600;
    color: {C_MUTED};
    font-size: 11px;
}}
QGroupBox::title {{
    subcontrol-origin: margin;
    subcontrol-position: top left;
    padding: 0 6px;
    left: 10px;
}}
QCheckBox {{
    spacing: 6px;
    color: {C_TEXT};
}}
QCheckBox::indicator {{
    width: 16px;
    height: 16px;
    border-radius: 3px;
    border: 1px solid {C_BORDER};
    background: {C_PANEL};
}}
QCheckBox::indicator:checked {{
    background: {C_ACCENT};
    border-color: {C_ACCENT};
}}
QProgressBar {{
    border: 1px solid {C_BORDER};
    border-radius: 4px;
    background: {C_PANEL};
    text-align: center;
    color: {C_TEXT};
    height: 14px;
}}
QProgressBar::chunk {{
    background: {C_ACCENT};
    border-radius: 3px;
}}
QSplitter::handle {{
    background: {C_BORDER};
    width: 1px;
}}
QLabel#section_title {{
    font-size: 11px;
    font-weight: bold;
    color: {C_MUTED};
    text-transform: uppercase;
    letter-spacing: 0.5px;
    padding: 2px 0;
}}
"""

DICTAMENES = ["", "Correcto", "Con observaciones", "Incompleto", "Fuera de plazo"]
DISCIPLINAS_LIST = ["No identificada", "Teatro", "Danza", "Música", "Artes Visuales", "Libro"]


# ─────────────────────────────────────────────────────────────
# SEMÁFORO WIDGET
# ─────────────────────────────────────────────────────────────

def status_color(status: str) -> str:
    return {
        engine.STATUS_OK:     C_OK,
        engine.STATUS_WARN:   C_WARN,
        engine.STATUS_FAIL:   C_FAIL,
        engine.STATUS_MANUAL: C_MANUAL,
        "reviewed":           C_REVIEWED,
    }.get(status, C_MUTED)


def status_icon(status: str) -> str:
    return {
        engine.STATUS_OK:     "✓",
        engine.STATUS_WARN:   "⚠",
        engine.STATUS_FAIL:   "✗",
        engine.STATUS_MANUAL: "○",
        "reviewed":           "✓",
    }.get(status, "?")


# ─────────────────────────────────────────────────────────────
# WORKERS
# ─────────────────────────────────────────────────────────────

class ProcessWorker(QThread):
    """Procesa una lista de PDFs en background."""
    progress    = pyqtSignal(int, int)          # (current, total)
    report_done = pyqtSignal(dict)              # report data
    finished    = pyqtSignal()
    error       = pyqtSignal(str, str)          # (path, message)

    def __init__(self, paths: list[str]):
        super().__init__()
        self._paths = paths

    def run(self):
        total = len(self._paths)
        for i, path in enumerate(self._paths):
            self.progress.emit(i + 1, total)
            try:
                report = engine.process_report(path)
                self.report_done.emit(report)
            except Exception as e:
                self.error.emit(path, str(e))
        self.finished.emit()


class SaveWorker(QThread):
    done    = pyqtSignal(str)
    errored = pyqtSignal(str)

    def __init__(self, path: str, data: dict):
        super().__init__()
        self._path = path
        self._data = data

    def run(self):
        try:
            class _Enc(json.JSONEncoder):
                def default(self, o):
                    try:
                        import numpy as np
                        if isinstance(o, np.integer): return int(o)
                        if isinstance(o, np.floating): return float(o) if not math.isnan(float(o)) else None
                    except ImportError:
                        pass
                    return super().default(o)

            text = json.dumps(self._data, cls=_Enc, indent=2, ensure_ascii=False)
            with open(self._path, "w", encoding="utf-8") as f:
                f.write(text)
            self.done.emit(self._path)
        except Exception as e:
            self.errored.emit(str(e))


# ─────────────────────────────────────────────────────────────
# VISOR PDF (reuso del Conciliador)
# ─────────────────────────────────────────────────────────────

class PDFViewer(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._pdf_bytes: bytes = b""
        self._total_pages: int = 0
        self._current_page: int = 0
        self._page_cache: dict = {}
        self._zoom = 1.0
        self._build()

    def _build(self):
        v = QVBoxLayout(self)
        v.setContentsMargins(0, 0, 0, 0)
        v.setSpacing(0)

        # Toolbar
        tb = QWidget()
        tb.setStyleSheet(f"background:{C_PANEL}; border-bottom:1px solid {C_BORDER};")
        tb_lay = QHBoxLayout(tb)
        tb_lay.setContentsMargins(8, 4, 8, 4)
        self.btn_prev = QPushButton("‹")
        self.btn_prev.setFixedSize(28, 28)
        self.btn_prev.clicked.connect(self._prev_page)
        self.lbl_page = QLabel("— / —")
        self.lbl_page.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_page.setStyleSheet(f"color:{C_MUTED}; font-size:12px; min-width:80px;")
        self.btn_next = QPushButton("›")
        self.btn_next.setFixedSize(28, 28)
        self.btn_next.clicked.connect(self._next_page)
        self.btn_zoom_in  = QPushButton("+")
        self.btn_zoom_in.setFixedSize(28, 28)
        self.btn_zoom_in.clicked.connect(self._zoom_in)
        self.btn_zoom_out = QPushButton("−")
        self.btn_zoom_out.setFixedSize(28, 28)
        self.btn_zoom_out.clicked.connect(self._zoom_out)
        tb_lay.addWidget(self.btn_prev)
        tb_lay.addWidget(self.lbl_page)
        tb_lay.addWidget(self.btn_next)
        tb_lay.addStretch()
        tb_lay.addWidget(self.btn_zoom_out)
        tb_lay.addWidget(self.btn_zoom_in)
        v.addWidget(tb)

        # Canvas
        self.scroll = QScrollArea()
        self.scroll.setWidgetResizable(True)
        self.scroll.setStyleSheet(f"background:{C_BG}; border:none;")
        self.lbl_img = QLabel()
        self.lbl_img.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_img.setStyleSheet(f"background:{C_BG};")
        self.scroll.setWidget(self.lbl_img)
        v.addWidget(self.scroll)

    def load_bytes(self, pdf_bytes: bytes):
        self._pdf_bytes = pdf_bytes
        self._page_cache.clear()
        self._current_page = 0
        if pdf_bytes:
            import fitz
            doc = fitz.open(stream=pdf_bytes, filetype="pdf")
            self._total_pages = doc.page_count
            doc.close()
        else:
            self._total_pages = 0
        self._render()

    def load_path(self, path: str):
        try:
            with open(path, "rb") as f:
                self.load_bytes(f.read())
        except Exception:
            self.load_bytes(b"")

    def _render(self):
        if not self._pdf_bytes or self._total_pages == 0:
            self.lbl_img.setText("Sin PDF cargado")
            self.lbl_page.setText("— / —")
            return
        self.lbl_page.setText(f"{self._current_page + 1} / {self._total_pages}")
        if self._current_page in self._page_cache:
            self.lbl_img.setPixmap(self._page_cache[self._current_page])
            return
        import fitz
        doc = fitz.open(stream=self._pdf_bytes, filetype="pdf")
        page = doc[self._current_page]
        mat = fitz.Matrix(2.0 * self._zoom, 2.0 * self._zoom)
        pix = page.get_pixmap(matrix=mat)
        doc.close()
        img = QImage(pix.samples, pix.width, pix.height, pix.stride, QImage.Format.Format_RGB888)
        pm = QPixmap.fromImage(img)
        self._page_cache[self._current_page] = pm
        self.lbl_img.setPixmap(pm)

    def _prev_page(self):
        if self._current_page > 0:
            self._current_page -= 1
            self._render()

    def _next_page(self):
        if self._current_page < self._total_pages - 1:
            self._current_page += 1
            self._render()

    def _zoom_in(self):
        self._zoom = min(self._zoom + 0.25, 3.0)
        self._page_cache.clear()
        self._render()

    def _zoom_out(self):
        self._zoom = max(self._zoom - 0.25, 0.5)
        self._page_cache.clear()
        self._render()


# ─────────────────────────────────────────────────────────────
# PANEL DE REVISIÓN (columna derecha)
# ─────────────────────────────────────────────────────────────

class NoScrollCombo(QComboBox):
    def wheelEvent(self, e): e.ignore()

class ReviewPanel(QWidget):
    """Panel con campos extraídos, checklist y observaciones."""
    changed = pyqtSignal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self._report: Optional[dict] = None
        self._build()

    def _build(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(12, 12, 12, 12)
        root.setSpacing(10)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("border:none;")
        inner = QWidget()
        v = QVBoxLayout(inner)
        v.setContentsMargins(0, 0, 8, 0)
        v.setSpacing(12)

        # ── Datos extraídos ──────────────────────────────────
        grp_datos = QGroupBox("Datos del informe")
        fl = QFormLayout(grp_datos)
        fl.setSpacing(8)
        fl.setLabelAlignment(Qt.AlignmentFlag.AlignRight)

        self.le_proyecto    = QLineEdit()
        self.le_erpi        = QLineEdit()
        self.cb_disciplina  = NoScrollCombo()
        self.cb_disciplina.addItems(DISCIPLINAS_LIST)
        self.le_etapa       = QLineEdit()
        self.le_num_informe = QLineEdit()
        self.le_periodo     = QLineEdit()
        self.le_fecha_inicio= QLineEdit()
        self.le_revisor     = QLineEdit()
        self.le_revisor.setPlaceholderText("Tu nombre")

        fl.addRow("Proyecto:",       self.le_proyecto)
        fl.addRow("ERPI:",           self.le_erpi)
        fl.addRow("Disciplina:",     self.cb_disciplina)
        fl.addRow("Etapa:",          self.le_etapa)
        fl.addRow("Nº Informe:",     self.le_num_informe)
        fl.addRow("Período:",        self.le_periodo)
        fl.addRow("Inicio recursos:",self.le_fecha_inicio)
        fl.addRow("Revisor:",        self.le_revisor)
        v.addWidget(grp_datos)

        # ── Checklist normativo ──────────────────────────────
        grp_check = QGroupBox("Checklist normativo")
        v_check = QVBoxLayout(grp_check)
        v_check.setSpacing(4)
        self._check_rows: list[QWidget] = []
        self._check_layout = v_check
        v.addWidget(grp_check)

        # ── Validación manual ────────────────────────────────
        grp_manual = QGroupBox("Validación manual")
        v_man = QVBoxLayout(grp_manual)
        self.chk_firma     = QCheckBox("Firma del representante legal presente")
        self.chk_id_oficial= QCheckBox("Identificación oficial adjunta")
        v_man.addWidget(self.chk_firma)
        v_man.addWidget(self.chk_id_oficial)
        v.addWidget(grp_manual)

        # ── Observaciones ────────────────────────────────────
        grp_obs = QGroupBox("Observaciones del revisor")
        v_obs = QVBoxLayout(grp_obs)
        self.te_obs = QTextEdit()
        self.te_obs.setPlaceholderText("Escribe aquí las observaciones...")
        self.te_obs.setMinimumHeight(100)
        v_obs.addWidget(self.te_obs)
        v.addWidget(grp_obs)

        # ── Dictamen ─────────────────────────────────────────
        grp_dict = QGroupBox("Dictamen")
        v_dict = QVBoxLayout(grp_dict)
        self.cb_dictamen = NoScrollCombo()
        self.cb_dictamen.addItems(DICTAMENES)
        v_dict.addWidget(self.cb_dictamen)

        row_btns = QHBoxLayout()
        self.btn_revisar = QPushButton("✓  Marcar como revisado")
        self.btn_revisar.setObjectName("btn_ok")
        self.btn_limpiar = QPushButton("Limpiar revisión")
        self.btn_limpiar.setObjectName("btn_secondary")
        row_btns.addWidget(self.btn_revisar)
        row_btns.addWidget(self.btn_limpiar)
        v_dict.addLayout(row_btns)
        v.addWidget(grp_dict)

        v.addStretch()
        scroll.setWidget(inner)
        root.addWidget(scroll)

        # Conectar señales
        for w in (self.le_proyecto, self.le_erpi, self.le_etapa,
                  self.le_num_informe, self.le_periodo, self.le_fecha_inicio, self.le_revisor):
            w.textChanged.connect(self.changed)
        self.cb_disciplina.currentTextChanged.connect(self.changed)
        self.chk_firma.stateChanged.connect(self.changed)
        self.chk_id_oficial.stateChanged.connect(self.changed)
        self.te_obs.textChanged.connect(self.changed)
        self.cb_dictamen.currentTextChanged.connect(self.changed)
        self.btn_revisar.clicked.connect(self._on_revisar)
        self.btn_limpiar.clicked.connect(self._on_limpiar)

    def _on_revisar(self):
        if self._report is not None:
            self._report["rev_revisado"] = True
            if not self._report["rev_dictamen"]:
                self._report["rev_dictamen"] = "Correcto"
                self.cb_dictamen.setCurrentText("Correcto")
            self.changed.emit()

    def _on_limpiar(self):
        if self._report is not None:
            self._report["rev_revisado"] = False
            self.changed.emit()

    def _build_checklist_rows(self, items: list):
        """Reconstruye los rows del checklist con los items actuales."""
        # Limpiar
        while self._check_rows:
            w = self._check_rows.pop()
            self._check_layout.removeWidget(w)
            w.deleteLater()

        for item in items:
            row = QWidget()
            row.setStyleSheet(f"background:transparent;")
            h = QHBoxLayout(row)
            h.setContentsMargins(0, 2, 0, 2)
            h.setSpacing(8)

            color = status_color(item.status)
            icon  = status_icon(item.status)

            lbl_icon = QLabel(icon)
            lbl_icon.setFixedWidth(18)
            lbl_icon.setStyleSheet(f"color:{color}; font-weight:bold; font-size:14px;")

            lbl_desc = QLabel(item.descripcion)
            lbl_desc.setStyleSheet(f"color:{C_TEXT}; font-size:12px;")
            lbl_desc.setWordWrap(True)

            lbl_detail = QLabel(item.detalle)
            lbl_detail.setStyleSheet(f"color:{C_MUTED}; font-size:11px;")
            lbl_detail.setWordWrap(True)

            col = QVBoxLayout()
            col.setSpacing(0)
            col.addWidget(lbl_desc)
            col.addWidget(lbl_detail)

            h.addWidget(lbl_icon)
            h.addLayout(col)
            h.addStretch()

            self._check_layout.addWidget(row)
            self._check_rows.append(row)

    def load_report(self, report: Optional[dict]):
        """Carga un informe en el panel."""
        self._report = report
        enabled = report is not None
        for w in (self.le_proyecto, self.le_erpi, self.cb_disciplina,
                  self.le_etapa, self.le_num_informe, self.le_periodo,
                  self.le_fecha_inicio, self.le_revisor, self.te_obs,
                  self.chk_firma, self.chk_id_oficial, self.cb_dictamen,
                  self.btn_revisar, self.btn_limpiar):
            w.setEnabled(enabled)

        if not enabled:
            return

        # Bloquear señales al cargar
        for w in (self.le_proyecto, self.le_erpi, self.le_etapa,
                  self.le_num_informe, self.le_periodo, self.le_fecha_inicio,
                  self.le_revisor, self.te_obs, self.cb_disciplina,
                  self.cb_dictamen, self.chk_firma, self.chk_id_oficial):
            w.blockSignals(True)

        self.le_proyecto.setText(report.get("rev_nombre_proyecto", ""))
        self.le_erpi.setText(report.get("rev_nombre_erpi", ""))
        disc = report.get("rev_disciplina", "No identificada")
        idx  = self.cb_disciplina.findText(disc)
        self.cb_disciplina.setCurrentIndex(idx if idx >= 0 else 0)
        self.le_etapa.setText(report.get("rev_etapa", ""))
        self.le_num_informe.setText(report.get("rev_numero_informe", ""))
        self.le_periodo.setText(report.get("rev_periodo", ""))
        self.le_fecha_inicio.setText(report.get("rev_fecha_inicio", ""))
        self.le_revisor.setText(report.get("rev_revisor", ""))
        self.te_obs.setPlainText(report.get("rev_observaciones", ""))
        self.chk_firma.setChecked(report.get("rev_firma", False))
        self.chk_id_oficial.setChecked(report.get("rev_id_oficial", False))
        dict_idx = self.cb_dictamen.findText(report.get("rev_dictamen", ""))
        self.cb_dictamen.setCurrentIndex(dict_idx if dict_idx >= 0 else 0)

        for w in (self.le_proyecto, self.le_erpi, self.le_etapa,
                  self.le_num_informe, self.le_periodo, self.le_fecha_inicio,
                  self.le_revisor, self.te_obs, self.cb_disciplina,
                  self.cb_dictamen, self.chk_firma, self.chk_id_oficial):
            w.blockSignals(False)

        # Checklist
        checklist = report.get("checklist", [])
        self._build_checklist_rows(checklist)

    def sync_to_report(self):
        """Copia los valores del panel al dict del informe."""
        if self._report is None:
            return
        r = self._report
        r["rev_nombre_proyecto"]  = self.le_proyecto.text().strip()
        r["rev_nombre_erpi"]      = self.le_erpi.text().strip()
        r["rev_disciplina"]       = self.cb_disciplina.currentText()
        r["rev_etapa"]            = self.le_etapa.text().strip()
        r["rev_numero_informe"]   = self.le_num_informe.text().strip()
        r["rev_periodo"]          = self.le_periodo.text().strip()
        r["rev_fecha_inicio"]     = self.le_fecha_inicio.text().strip()
        r["rev_revisor"]          = self.le_revisor.text().strip()
        r["rev_observaciones"]    = self.te_obs.toPlainText().strip()
        r["rev_firma"]            = self.chk_firma.isChecked()
        r["rev_id_oficial"]       = self.chk_id_oficial.isChecked()
        r["rev_dictamen"]         = self.cb_dictamen.currentText()


# ─────────────────────────────────────────────────────────────
# ITEM DE LISTA
# ─────────────────────────────────────────────────────────────

STATUS_ORDER = {engine.STATUS_FAIL: 0, engine.STATUS_WARN: 1,
                "reviewed": 2, engine.STATUS_OK: 3}


def report_display_status(report: dict) -> str:
    if report.get("rev_revisado"):
        return "reviewed"
    return report.get("status", engine.STATUS_FAIL)


def make_list_item(report: dict) -> QListWidgetItem:
    status = report_display_status(report)
    color  = status_color(status)
    icon   = status_icon(status)
    fname  = report.get("filename", "?")
    disc   = report.get("rev_disciplina", "?")
    text   = f"{icon}  {fname}\n      {disc}"
    item   = QListWidgetItem(text)
    item.setForeground(QColor(color if status == engine.STATUS_FAIL else C_TEXT))
    item.setData(Qt.ItemDataRole.UserRole, id(report))
    return item


# ─────────────────────────────────────────────────────────────
# VENTANA PRINCIPAL
# ─────────────────────────────────────────────────────────────

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Revisor de Informes Semestrales — EFIARTES")
        self.resize(1500, 900)
        self._reports: list[dict] = []
        self._current_idx: Optional[int] = None
        self._session_file: str = ""
        self._save_worker = None
        self._process_worker = None
        self._unsaved = False
        self._build()

    # ── UI ─────────────────────────────────────────────────

    def _build(self):
        self.setStyleSheet(GLOBAL_STYLE)

        # Barra de estado
        self.status = QStatusBar()
        self.status.setStyleSheet(f"background:{C_PANEL}; color:{C_MUTED}; font-size:12px;"
                                  f" border-top:1px solid {C_BORDER};")
        self.setStatusBar(self.status)

        # ── Banner superior ───────────────────────────────
        banner = QWidget()
        banner.setFixedHeight(52)
        banner.setStyleSheet(f"background: qlineargradient(x1:0,y1:0,x2:1,y2:0,"
                             f"stop:0 #0D1F3C, stop:1 {C_BG});"
                             f"border-bottom:1px solid {C_BORDER};")
        b_lay = QHBoxLayout(banner)
        b_lay.setContentsMargins(12, 0, 12, 0)
        b_lay.setSpacing(8)

        lbl_title = QLabel("EFIARTES — Revisor de Informes Semestrales")
        lbl_title.setStyleSheet(f"color:{C_TEXT}; font-size:15px; font-weight:700;")

        self.btn_save = QPushButton("💾  Guardar sesión")
        self.btn_save.setFixedHeight(34)
        self.btn_save.clicked.connect(self._save_session)

        self.btn_export = QPushButton("📊  Exportar Excel")
        self.btn_export.setObjectName("btn_secondary")
        self.btn_export.setFixedHeight(34)
        self.btn_export.clicked.connect(self._export_excel)

        b_lay.addWidget(lbl_title)
        b_lay.addStretch()
        b_lay.addWidget(self.btn_save)
        b_lay.addWidget(self.btn_export)

        # Barra de progreso
        self.progress_bar = QProgressBar()
        self.progress_bar.setFixedHeight(3)
        self.progress_bar.setTextVisible(False)
        self.progress_bar.setVisible(False)
        self.progress_bar.setStyleSheet(f"QProgressBar{{border:none;background:{C_PANEL};}}"
                                        f"QProgressBar::chunk{{background:{C_ACCENT};}}")

        # ── Splitter principal ───────────────────────────
        splitter = QSplitter(Qt.Orientation.Horizontal)
        splitter.setHandleWidth(2)

        # Columna izquierda — Bandeja
        left = self._build_left()
        splitter.addWidget(left)

        # Centro — Visor PDF
        self.pdf_viewer = PDFViewer()
        splitter.addWidget(self.pdf_viewer)

        # Derecha — Panel de revisión
        self.review_panel = ReviewPanel()
        self.review_panel.setMinimumWidth(380)
        self.review_panel.changed.connect(self._on_review_changed)
        splitter.addWidget(self.review_panel)

        splitter.setSizes([300, 700, 420])

        # Root layout
        central = QWidget()
        root = QVBoxLayout(central)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)
        root.addWidget(banner)
        root.addWidget(self.progress_bar)
        root.addWidget(splitter)
        self.setCentralWidget(central)

        # Menú
        self._build_menu()

        # Estado inicial
        self.review_panel.load_report(None)

    def _build_left(self) -> QWidget:
        w = QWidget()
        w.setMinimumWidth(260)
        w.setMaximumWidth(380)
        w.setStyleSheet(f"background:{C_SIDEBAR}; border-right:1px solid {C_BORDER};")
        v = QVBoxLayout(w)
        v.setContentsMargins(10, 10, 10, 10)
        v.setSpacing(8)

        # KPIs rápidos
        self.lbl_total    = self._kpi_label("Total informes", "0")
        self.lbl_ok       = self._kpi_label("Completos", "0", C_OK)
        self.lbl_warn     = self._kpi_label("Con advertencias", "0", C_WARN)
        self.lbl_fail     = self._kpi_label("Incompletos", "0", C_FAIL)
        self.lbl_reviewed = self._kpi_label("Revisados", "0", C_ACCENT)

        kpi_row1 = QHBoxLayout()
        kpi_row1.addWidget(self.lbl_total)
        kpi_row1.addWidget(self.lbl_reviewed)
        kpi_row2 = QHBoxLayout()
        kpi_row2.addWidget(self.lbl_ok)
        kpi_row2.addWidget(self.lbl_warn)
        kpi_row2.addWidget(self.lbl_fail)
        v.addLayout(kpi_row1)
        v.addLayout(kpi_row2)

        # Filtro
        self.le_filter = QLineEdit()
        self.le_filter.setPlaceholderText("🔍 Filtrar por nombre, ERPI, disciplina…")
        self.le_filter.textChanged.connect(self._apply_filter)
        v.addWidget(self.le_filter)

        self.cb_filter_status = NoScrollCombo()
        self.cb_filter_status.addItems(["Todos los estados", "Incompleto ✗", "Advertencia ⚠", "Completo ✓", "Revisado ✓"])
        self.cb_filter_status.currentTextChanged.connect(self._apply_filter)
        v.addWidget(self.cb_filter_status)

        # Lista
        self.list_reports = QListWidget()
        self.list_reports.setSpacing(2)
        self.list_reports.currentRowChanged.connect(self._on_report_selected)
        v.addWidget(self.list_reports)

        # Botones
        btn_add = QPushButton("+ Agregar PDF(s)")
        btn_add.clicked.connect(self._add_pdfs)
        btn_remove = QPushButton("Quitar seleccionado")
        btn_remove.setObjectName("btn_secondary")
        btn_remove.clicked.connect(self._remove_selected)
        btn_load_session = QPushButton("Abrir sesión (.efias)")
        btn_load_session.setObjectName("btn_secondary")
        btn_load_session.clicked.connect(self._load_session)

        v.addWidget(btn_add)
        v.addWidget(btn_remove)
        v.addWidget(btn_load_session)
        return w

    def _kpi_label(self, title: str, value: str, color: str = C_MUTED) -> QWidget:
        w = QWidget()
        w.setStyleSheet(f"background:{C_PANEL}; border-radius:6px; border:1px solid {C_BORDER};")
        v = QVBoxLayout(w)
        v.setContentsMargins(8, 6, 8, 6)
        v.setSpacing(0)
        lbl_v = QLabel(value)
        lbl_v.setStyleSheet(f"color:{color}; font-size:20px; font-weight:700;")
        lbl_v.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lbl_t = QLabel(title)
        lbl_t.setStyleSheet(f"color:{C_MUTED}; font-size:10px;")
        lbl_t.setAlignment(Qt.AlignmentFlag.AlignCenter)
        v.addWidget(lbl_v)
        v.addWidget(lbl_t)
        setattr(self, f"_kpi_{title.replace(' ','_').lower()}_label", lbl_v)
        return w

    def _build_menu(self):
        mb = self.menuBar()
        mb.setStyleSheet(f"background:{C_PANEL}; color:{C_TEXT}; border-bottom:1px solid {C_BORDER};")

        file_m = mb.addMenu("Archivo")
        file_m.addAction("Agregar PDF(s)…", self._add_pdfs)
        file_m.addSeparator()
        file_m.addAction("Abrir sesión (.efias)…", self._load_session)
        file_m.addAction("Guardar sesión", self._save_session)
        file_m.addAction("Guardar sesión como…", self._save_session_as)
        file_m.addSeparator()
        file_m.addAction("Exportar a Excel…", self._export_excel)
        file_m.addSeparator()
        file_m.addAction("Salir", self.close)

    # ── Lógica de PDF ───────────────────────────────────────

    def _add_pdfs(self):
        paths, _ = QFileDialog.getOpenFileNames(
            self, "Seleccionar informes PDF", "",
            "Archivos PDF (*.pdf);;Todos los archivos (*)"
        )
        if not paths:
            return
        # Evitar duplicados
        existing = {r["path"] for r in self._reports}
        new_paths = [p for p in paths if p not in existing]
        if not new_paths:
            self.status.showMessage("Todos los archivos seleccionados ya estaban cargados.", 3000)
            return
        self._process_pdfs(new_paths)

    def _process_pdfs(self, paths: list[str]):
        self.progress_bar.setVisible(True)
        self.progress_bar.setMaximum(len(paths))
        self.progress_bar.setValue(0)
        self.status.showMessage(f"Procesando {len(paths)} archivo(s)…")

        worker = ProcessWorker(paths)
        worker.progress.connect(lambda c, t: self.progress_bar.setValue(c))
        worker.report_done.connect(self._on_report_ready)
        worker.error.connect(lambda p, e: self.status.showMessage(f"Error: {os.path.basename(p)} — {e}", 5000))
        worker.finished.connect(self._on_processing_done)
        self._process_worker = worker
        worker.start()

    def _on_report_ready(self, report: dict):
        self._reports.append(report)
        item = make_list_item(report)
        self.list_reports.addItem(item)
        self._update_kpis()
        self._unsaved = True

    def _on_processing_done(self):
        self.progress_bar.setVisible(False)
        n = len(self._reports)
        self.status.showMessage(f"{n} informe(s) cargado(s). Selecciona uno para revisarlo.", 4000)

    def _remove_selected(self):
        row = self.list_reports.currentRow()
        if row < 0:
            return
        # Encontrar el reporte real
        item = self.list_reports.item(row)
        report_id = item.data(Qt.ItemDataRole.UserRole)
        self._reports = [r for r in self._reports if id(r) != report_id]
        self.list_reports.takeItem(row)
        self._current_idx = None
        self.review_panel.load_report(None)
        self.pdf_viewer.load_bytes(b"")
        self._update_kpis()

    # ── Selección de informe ────────────────────────────────

    def _on_report_selected(self, row: int):
        if row < 0 or row >= len(self._reports):
            return
        # Guardar cambios del informe anterior
        self.review_panel.sync_to_report()

        self._current_idx = row
        report = self._reports[row]
        self.review_panel.load_report(report)
        self.pdf_viewer.load_path(report["path"])
        self.status.showMessage(f"Revisando: {report['filename']}", 0)

    def _on_review_changed(self):
        """Sincroniza el panel al informe y actualiza el item de la lista."""
        self.review_panel.sync_to_report()
        if self._current_idx is not None and self._current_idx < len(self._reports):
            report = self._reports[self._current_idx]
            item = self.list_reports.item(self._current_idx)
            if item:
                new_item = make_list_item(report)
                item.setText(new_item.text())
                item.setForeground(new_item.foreground())
        self._update_kpis()
        self._unsaved = True

    def _apply_filter(self):
        query = self.le_filter.text().lower()
        status_filter = self.cb_filter_status.currentText()
        for i, report in enumerate(self._reports):
            item = self.list_reports.item(i)
            if not item:
                continue
            # Filtro de texto
            text_match = (
                query in report.get("filename", "").lower() or
                query in report.get("rev_nombre_proyecto", "").lower() or
                query in report.get("rev_nombre_erpi", "").lower() or
                query in report.get("rev_disciplina", "").lower()
            )
            # Filtro de estado
            st = report_display_status(report)
            status_match = (
                status_filter == "Todos los estados" or
                (status_filter.startswith("Incompleto") and st == engine.STATUS_FAIL) or
                (status_filter.startswith("Advertencia") and st == engine.STATUS_WARN) or
                (status_filter.startswith("Completo") and st == engine.STATUS_OK) or
                (status_filter.startswith("Revisado") and st == "reviewed")
            )
            item.setHidden(not (text_match and status_match))

    def _update_kpis(self):
        n = len(self._reports)
        reviewed = sum(1 for r in self._reports if r.get("rev_revisado"))
        ok   = sum(1 for r in self._reports if report_display_status(r) == engine.STATUS_OK)
        warn = sum(1 for r in self._reports if report_display_status(r) == engine.STATUS_WARN)
        fail = sum(1 for r in self._reports if report_display_status(r) == engine.STATUS_FAIL)

        self._update_kpi_val("Total informes", str(n))
        self._update_kpi_val("Revisados", str(reviewed))
        self._update_kpi_val("Completos", str(ok))
        self._update_kpi_val("Con advertencias", str(warn))
        self._update_kpi_val("Incompletos", str(fail))

    def _update_kpi_val(self, title: str, val: str):
        attr = f"_kpi_{title.replace(' ','_').lower()}_label"
        lbl = getattr(self, attr, None)
        if lbl:
            lbl.setText(val)

    # ── Guardar / Cargar sesión ─────────────────────────────

    def _session_to_dict(self) -> dict:
        """Serializa todos los informes para guardar en .efias."""
        reports_data = []
        for r in self._reports:
            rd = {k: v for k, v in r.items() if k not in ("checklist", "text", "fields", "texto_completo")}
            # Guardar campos que necesitamos para reconstruir la UI
            rd["checklist"] = [
                {"id": c.id, "descripcion": c.descripcion, "status": c.status, "detalle": c.detalle}
                for c in r.get("checklist", [])
            ]
            reports_data.append(rd)
        return {"version": "1.0", "reports": reports_data}

    def _save_session(self):
        if self._session_file:
            self._do_save(self._session_file)
        else:
            self._save_session_as()

    def _save_session_as(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Guardar sesión EFIARTES", "",
            "Sesión EFIARTES (*.efias);;Todos los archivos (*)"
        )
        if not path:
            return
        if not path.endswith(".efias"):
            path += ".efias"
        self._do_save(path)

    def _do_save(self, path: str):
        self.review_panel.sync_to_report()
        self.btn_save.setEnabled(False)
        self.status.showMessage("Guardando sesión…")
        data = self._session_to_dict()

        def on_done(p):
            self._session_file = p
            self.btn_save.setEnabled(True)
            self._unsaved = False
            self.status.showMessage(f"Sesión guardada: {p}", 4000)

        def on_err(msg):
            self.btn_save.setEnabled(True)
            QMessageBox.critical(self, "Error al guardar", f"No se pudo guardar:\n{msg}")

        w = SaveWorker(path, data)
        w.done.connect(on_done)
        w.errored.connect(on_err)
        self._save_worker = w
        w.start()

    def _load_session(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Abrir sesión EFIARTES", "",
            "Sesión EFIARTES (*.efias);;Todos los archivos (*)"
        )
        if not path:
            return
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            self._reports.clear()
            self.list_reports.clear()
            for rd in data.get("reports", []):
                # Reconstruir checklist como objetos
                checklist_raw = rd.pop("checklist", [])
                checklist = []
                for c in checklist_raw:
                    ci = engine.CheckItem(
                        id=c.get("id", ""),
                        descripcion=c.get("descripcion", ""),
                        status=c.get("status", engine.STATUS_MANUAL),
                        detalle=c.get("detalle", ""),
                    )
                    checklist.append(ci)
                rd["checklist"] = checklist
                rd.setdefault("text", "")
                rd.setdefault("fields", {})
                self._reports.append(rd)
                item = make_list_item(rd)
                self.list_reports.addItem(item)
            self._session_file = path
            self._unsaved = False
            self._update_kpis()
            self.status.showMessage(f"Sesión cargada: {path} — {len(self._reports)} informes.", 4000)
        except Exception as e:
            QMessageBox.critical(self, "Error al abrir", f"No se pudo abrir la sesión:\n{e}")

    # ── Exportar Excel ──────────────────────────────────────

    def _export_excel(self):
        if not self._reports:
            self.status.showMessage("No hay informes cargados.", 3000)
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Exportar reporte Excel", "Revision_EFIARTES.xlsx",
            "Excel (*.xlsx);;Todos los archivos (*)"
        )
        if not path:
            return
        self.review_panel.sync_to_report()
        try:
            self._write_excel(path)
            self.status.showMessage(f"Excel exportado: {path}", 4000)
        except Exception as e:
            QMessageBox.critical(self, "Error al exportar", f"No se pudo exportar:\n{e}")

    def _write_excel(self, path: str):
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Revisión EFIARTES"

        # Colores
        hdr_fill = PatternFill("solid", fgColor="0D1F3C")
        ok_fill  = PatternFill("solid", fgColor="1A3A22")
        warn_fill= PatternFill("solid", fgColor="3A2E10")
        fail_fill= PatternFill("solid", fgColor="3A1010")
        rev_fill = PatternFill("solid", fgColor="1A2A3A")
        hdr_font = Font(color="E6EDF3", bold=True, name="Calibri", size=11)
        thin = Border(
            left=Side(style="thin", color="2D3748"),
            right=Side(style="thin", color="2D3748"),
            top=Side(style="thin", color="2D3748"),
            bottom=Side(style="thin", color="2D3748"),
        )

        headers = [
            "Archivo", "Disciplina", "Nombre del Proyecto", "ERPI",
            "Etapa", "Nº Informe", "Período", "Fecha Inicio Recursos",
            "Firma", "ID Oficial", "R1", "R2", "R3", "R4", "R5", "R6", "R7",
            "Semáforo Auto", "Revisado por", "Dictamen", "Observaciones", "Revisado"
        ]
        ws.append(headers)
        for col, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin

        status_map = {
            engine.STATUS_OK: "✓ Completo",
            engine.STATUS_WARN: "⚠ Advertencia",
            engine.STATUS_FAIL: "✗ Incompleto",
        }

        for r in self._reports:
            checklist = r.get("checklist", [])
            check_vals = {c.id: status_icon(c.status) for c in checklist}
            st = report_display_status(r)
            row_data = [
                r.get("filename", ""),
                r.get("rev_disciplina", ""),
                r.get("rev_nombre_proyecto", ""),
                r.get("rev_nombre_erpi", ""),
                r.get("rev_etapa", ""),
                r.get("rev_numero_informe", ""),
                r.get("rev_periodo", ""),
                r.get("rev_fecha_inicio", ""),
                "Sí" if r.get("rev_firma") else "No",
                "Sí" if r.get("rev_id_oficial") else "No",
                check_vals.get("R1", "?"),
                check_vals.get("R2", "?"),
                check_vals.get("R3", "?"),
                check_vals.get("R4", "?"),
                check_vals.get("R5", "?"),
                check_vals.get("R6", "?"),
                check_vals.get("R7", "?"),
                status_map.get(st, st),
                r.get("rev_revisor", ""),
                r.get("rev_dictamen", ""),
                r.get("rev_observaciones", ""),
                "Sí" if r.get("rev_revisado") else "No",
            ]
            ws.append(row_data)
            row_num = ws.max_row
            fill = (rev_fill if r.get("rev_revisado") else
                    fail_fill if st == engine.STATUS_FAIL else
                    warn_fill if st == engine.STATUS_WARN else ok_fill)
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col)
                cell.fill = fill
                cell.font = Font(color="E6EDF3", name="Calibri", size=10)
                cell.border = thin
                cell.alignment = Alignment(vertical="center", wrap_text=True)

        # Anchos de columna
        col_widths = [35, 18, 40, 35, 18, 20, 15, 22, 8, 10, 5, 5, 5, 5, 5, 5, 5,
                      18, 20, 18, 50, 10]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.row_dimensions[1].height = 32
        ws.freeze_panes = "A2"

        wb.save(path)

    # ── Cierre ──────────────────────────────────────────────

    def closeEvent(self, event):
        if self._unsaved and self._reports:
            reply = QMessageBox.question(
                self, "Cambios sin guardar",
                "Tienes cambios sin guardar. ¿Deseas guardar antes de salir?",
                QMessageBox.StandardButton.Save |
                QMessageBox.StandardButton.Discard |
                QMessageBox.StandardButton.Cancel
            )
            if reply == QMessageBox.StandardButton.Save:
                self._save_session()
                event.accept()
            elif reply == QMessageBox.StandardButton.Cancel:
                event.ignore()
            else:
                event.accept()
        else:
            event.accept()


# ─────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────

if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setApplicationName("Revisor EFIARTES")
    app.setStyle("Fusion")
    win = MainWindow()
    win.show()
    sys.exit(app.exec())
